
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import re
from io import BytesIO

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

full_bank_txns = None
full_book_txns = None

def extract_numbers_with_tail(text):
    if text is None:
        return set()
    tokens = re.findall(r'\d+', str(text))
    result = set(tokens)
    for token in tokens:
        if len(token) >= 4:
            result.add(token[-4:])
    return result

def extract_transactions(ws):
    bank_txns = []
    book_txns = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        bank_amount = row[3].value
        bank_id = row[1].value
        bank_date = row[0].value
        bank_dr_cr = row[4].value
        bank_narration = row[2].value

        book_amount = row[5].value
        book_id = row[7].value
        book_date = row[8].value
        book_col_j = row[9].value
        book_description = row[13].value
        book_dr_cr = row[15].value

        if bank_amount is not None:
            bank_txns.append({
                'row_idx': row_idx,
                'row': row,
                'date': bank_date,
                'id': bank_id,
                'amount': bank_amount,
                'dr_cr': bank_dr_cr,
                'narration': bank_narration,
                'cell': row[3],
                'matched': False,
                'narration_matched': False
            })

        if book_amount is not None:
            book_txns.append({
                'row_idx': row_idx,
                'row': row,
                'date': book_date,
                'id': book_id,
                'col_j_val': book_col_j,
                'amount': book_amount,
                'dr_cr': book_dr_cr,
                'description': book_description,
                'cell': row[5],
                'matched': False,
                'reversal': False,
                'narration_matched': False
            })
    return bank_txns, book_txns

def colorize_cells(bank_txns, book_txns):
    for txn in bank_txns:
        txn['cell'].fill = green_fill if txn['matched'] else red_fill
    for txn in book_txns:
        txn['cell'].fill = green_fill if txn['matched'] else red_fill

def process_amount_only(wb):
    ws = wb.active
    bank_txns, book_txns = extract_transactions(ws)
    for btxn in bank_txns:
        for bktxn in book_txns:
            if not btxn['matched'] and not bktxn['matched']:
                if btxn['amount'] == bktxn['amount']:
                    btxn['matched'] = True
                    bktxn['matched'] = True
                    break
    colorize_cells(bank_txns, book_txns)
    return wb

def process_drcr(wb):
    ws = wb.active
    bank_txns, book_txns = extract_transactions(ws)
    for btxn in bank_txns:
        for bktxn in book_txns:
            if not btxn['matched'] and not bktxn['matched']:
                if btxn['amount'] == bktxn['amount'] and (
                    (btxn['dr_cr'] == "DR" and bktxn['dr_cr'] == "CR") or
                    (btxn['dr_cr'] == "CR" and bktxn['dr_cr'] == "DR")
                ):
                    btxn['matched'] = True
                    bktxn['matched'] = True
                    break
    colorize_cells(bank_txns, book_txns)
    return wb

def process_date(wb):
    ws = wb.active
    bank_txns, book_txns = extract_transactions(ws)
    for btxn in bank_txns:
        for bktxn in book_txns:
            if not btxn['matched'] and not bktxn['matched']:
                if btxn['amount'] == bktxn['amount'] and (
                    (btxn['dr_cr'] == "DR" and bktxn['dr_cr'] == "CR") or
                    (btxn['dr_cr'] == "CR" and bktxn['dr_cr'] == "DR")
                ):
                    if btxn['date'] and bktxn['date']:
                        if abs((btxn['date'] - bktxn['date']).days) <= 10:
                            btxn['matched'] = True
                            bktxn['matched'] = True
                            break
    colorize_cells(bank_txns, book_txns)
    return wb

def process_full(wb):
    global full_bank_txns, full_book_txns
    ws = wb.active
    bank_txns, book_txns = extract_transactions(ws)
    standard_matched_pairs = []
    for btxn in bank_txns:
        for bktxn in book_txns:
            if not btxn['matched'] and not bktxn['matched']:
                if btxn['amount'] == bktxn['amount'] and (
                    (btxn['dr_cr'] == "DR" and bktxn['dr_cr'] == "CR") or
                    (btxn['dr_cr'] == "CR" and bktxn['dr_cr'] == "DR")
                ):
                    if btxn['date'] and bktxn['date']:
                        if abs((btxn['date'] - bktxn['date']).days) <= 10:
                            btxn['matched'] = True
                            bktxn['matched'] = True
                            standard_matched_pairs.append((btxn, bktxn))
                            break
    for btxn, bktxn in standard_matched_pairs:
        bank_tokens = extract_numbers_with_tail(btxn['narration'])
        book_tokens = extract_numbers_with_tail(bktxn['description'])
        if bank_tokens.intersection(book_tokens):
            btxn['narration_matched'] = True
            bktxn['narration_matched'] = True
    colorize_cells(bank_txns, book_txns)
    full_bank_txns = bank_txns
    full_book_txns = book_txns
    return wb

def process_uploaded_file(file, option):
    wb = load_workbook(filename=BytesIO(file.getvalue()))
    if option == "Matched Amounts (Amount Only)":
        return process_amount_only(wb)
    elif option == "DRCR/CRDR (Amount + DR/CR)":
        return process_drcr(wb)
    elif option == "Date (Amount + DR/CR + Date)":
        return process_date(wb)
    elif option == "Check Narration (All Conditions)":
        return process_full(wb)
    else:
        return wb

st.set_page_config(page_title="Bank Reconciliation", layout="centered")
st.title("Bank Reconciliation App")

uploaded_file = st.file_uploader("Upload Excel File", type="xlsx")

if uploaded_file:
    st.success("File uploaded. Choose a reconciliation step below.")
    option = st.selectbox("Select Stage", [
        "Matched Amounts (Amount Only)",
        "DRCR/CRDR (Amount + DR/CR)",
        "Date (Amount + DR/CR + Date)",
        "Check Narration (All Conditions)"
    ])

    if st.button("Run Reconciliation"):
        st.info(f"Processing: {option}")
        processed_wb = process_uploaded_file(uploaded_file, option)
        output = BytesIO()
        processed_wb.save(output)
        st.success("Processing complete. Click below to download.")
        st.download_button(
            label="Download Processed Excel",
            data=output.getvalue(),
            file_name="Processed_BRS.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.warning("Upload a valid Excel file to continue.")
